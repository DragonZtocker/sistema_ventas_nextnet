
from flask import Blueprint, render_template, request, send_file
from sqlalchemy import func, extract
from models import db, Venta
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

bp = Blueprint("reportes", __name__, url_prefix="/reportes")

def parse_date(s: str):
    from datetime import datetime
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

def apply_filters(query, f_inicio, f_fin, consultor, anio):
    if f_inicio:
        query = query.filter(Venta.fecha >= f_inicio)
    if f_fin:
        query = query.filter(Venta.fecha <= f_fin)
    if consultor:
        query = query.filter(Venta.nombre_consultor.ilike(f"%{consultor}%"))
    if anio:
        try:
            query = query.filter(extract("year", Venta.fecha) == int(anio))
        except Exception:
            pass
    return query

@bp.route("/", methods=["GET"])
def dashboard():
    f_inicio = parse_date(request.args.get("inicio", ""))
    f_fin = parse_date(request.args.get("fin", ""))
    consultor = (request.args.get("consultor") or "").strip()
    anio = (request.args.get("anio") or "").strip()

    base_q = apply_filters(db.session.query(Venta), f_inicio, f_fin, consultor, anio)
    ventas = base_q.order_by(Venta.fecha.desc()).all()

    def sum_col(col):
        return apply_filters(db.session.query(func.coalesce(func.sum(col), 0)), f_inicio, f_fin, consultor, anio).scalar() or 0

    totals = {
        "mrc_nuevo": sum_col(Venta.mrc_nuevo),
        "fcv_nuevo": sum_col(Venta.fcv_nuevo),
        "pago_unico": sum_col(Venta.pago_unico),
        "fcv_renovado": sum_col(Venta.fcv_renovado),
        "mrc_final": sum_col(Venta.mrc_final),
        "variacion": sum_col(Venta.variacion),
    }

    # Ranking por consultor (Σ MRC Nuevo y Σ FCV Nuevo), ordenado por Σ MRC Nuevo
    ranking_q = db.session.query(
        Venta.nombre_consultor,
        func.count(Venta.id).label("num"),
        func.coalesce(func.sum(Venta.mrc_nuevo), 0).label("mrc_nuevo_sum"),
        func.coalesce(func.sum(Venta.fcv_nuevo), 0).label("fcv_nuevo_sum"),
    )
    ranking_q = apply_filters(ranking_q, f_inicio, f_fin, consultor, anio)
    ranking = (
        ranking_q.group_by(Venta.nombre_consultor)
        .order_by(func.coalesce(func.sum(Venta.mrc_nuevo), 0).desc())
        .all()
    )

    return render_template("reportes.html", ventas=ventas, totals=totals, ranking=ranking)

@bp.route("/excel")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, numbers

    f_inicio = parse_date(request.args.get("inicio", ""))
    f_fin = parse_date(request.args.get("fin", ""))
    consultor = (request.args.get("consultor") or "").strip()
    anio = (request.args.get("anio") or "").strip()

    q = apply_filters(Venta.query, f_inicio, f_fin, consultor, anio)
    rows = q.order_by(Venta.fecha.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = [
        "Fecha","Periodo","Sector","Consultor","Cliente","Origen",
        "Cotización Odoo","Licitación","Tipo Servicio","Plazo","Tipo Venta",
        "FCV Nuevo","MRC Nuevo","FCV Renovado","MRC Inicial","MRC Final","Variación","Pago Único"
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for v in rows:
        ws.append([
            v.fecha.isoformat(),
            v.periodo,
            v.sector,
            v.nombre_consultor,
            v.nombre_cliente,
            v.origen_venta,
            v.n_cotizacion_odoo or "",
            v.n_licitacion or "",
            v.tipo_servicio,
            v.plazo_contrato,
            v.tipo_venta,
            float(v.fcv_nuevo or 0),
            float(v.mrc_nuevo or 0),
            float(v.fcv_renovado or 0),
            float(v.mrc_inicial or 0),
            float(v.mrc_final or 0),
            float(v.variacion or 0),
            float(v.pago_unico or 0),
        ])

    money_cols = [12,13,14,15,16,17,18]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for idx in money_cols:
            row[idx-1].number_format = numbers.FORMAT_NUMBER_00

    # Autosize
    for col in range(1, len(headers)+1):
        letter = get_column_letter(col)
        maxlen = 0
        for cell in ws[letter]:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > maxlen:
                maxlen = len(val)
        ws.column_dimensions[letter].width = min(max(10, maxlen + 2), 50)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="reporte_ventas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@bp.route("/pdf")
def export_pdf():
    f_inicio = parse_date(request.args.get("inicio", ""))
    f_fin = parse_date(request.args.get("fin", ""))
    consultor = (request.args.get("consultor") or "").strip()
    anio = (request.args.get("anio") or "").strip()

    q = apply_filters(Venta.query, f_inicio, f_fin, consultor, anio)
    rows = q.order_by(Venta.fecha.desc()).all()

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    width, height = landscape(A4)
    y = height - 30
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30, y, "Reporte de Ventas")
    y -= 20
    c.setFont("Helvetica", 8)
    headers = ["Fecha","Consultor","Cliente","Tipo","Plazo","MRC Inicial","MRC Final","Variación","FCV Nuevo","FCV Renovado","Pago Único"]
    x_positions = [30, 100, 220, 380, 430, 480, 540, 600, 660, 740, 820]
    for i,h in enumerate(headers):
        c.drawString(x_positions[i], y, h)
    y -= 14
    for v in rows:
        values = [
            v.fecha.isoformat(),
            (v.nombre_consultor or "")[:18],
            (v.nombre_cliente or "")[:20],
            v.tipo_venta or "",
            str(v.plazo_contrato or ""),
            f"{float(v.mrc_inicial or 0):.2f}",
            f"{float(v.mrc_final or 0):.2f}",
            f"{float(v.variacion or 0):.2f}",
            f"{float(v.fcv_nuevo or 0):.2f}",
            f"{float(v.fcv_renovado or 0):.2f}",
            f"{float(v.pago_unico or 0):.2f}",
        ]
        for i,val in enumerate(values):
            c.drawString(x_positions[i], y, str(val))
        y -= 12
        if y < 40:
            c.showPage(); y = height - 40
            c.setFont("Helvetica", 8)
    c.save()
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="reporte_ventas.pdf", mimetype="application/pdf")
